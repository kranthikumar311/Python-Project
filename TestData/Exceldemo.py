import openpyxl
book = openpyxl.load_workbook("F:\\SELENIUM USING PYTHON\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=1)
# print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)









